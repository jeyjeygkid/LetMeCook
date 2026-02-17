import os
import io
import re
import zipfile
from datetime import datetime, timezone

import openpyxl
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, abort
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

APP_NAME = "Let Me Cook!"
APP_TAGLINE = "Running Dinner"

APP_SECRET = os.environ.get("APP_SECRET", "developer")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Batcom")

EVENT_PIN = (os.environ.get("EVENT_PIN", "1337") or "").strip()
BASE_URL = os.environ.get("BASE_URL", "http://127.0.0.1:5000")

DATABASE_URL = (os.environ.get("DATABASE_URL", "") or "").strip()

STORAGE_BACKEND = (os.environ.get("STORAGE_BACKEND", "db") or "db").strip().lower()
S3_BUCKET = (os.environ.get("S3_BUCKET", "") or "").strip()
S3_ENDPOINT_URL = (os.environ.get("S3_ENDPOINT_URL", "") or "").strip()

COURSE_OPTIONS = ["Vorspeise", "Hauptgang", "Nachtisch", "Egal"]
DIET_OPTIONS = ["Nichts", "Vegetarisch", "Vegan"]

XLSX_TEMPLATE_PATH = os.environ.get(
    "XLSX_TEMPLATE_PATH",
    os.path.join(os.path.dirname(__file__), "RudiRockt_Table1_template.xlsx")
)

app = Flask(__name__)
app.secret_key = APP_SECRET

# DB config: Postgres if provided; else SQLite file for local testing.
if DATABASE_URL:
    db_uri = DATABASE_URL
else:
    os.makedirs(app.instance_path, exist_ok=True)
    db_uri = "sqlite:///" + os.path.join(app.instance_path, "app.db")

app.config["SQLALCHEMY_DATABASE_URI"] = db_uri
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)

# ---------------- Models ----------------
class User(db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String, unique=True, nullable=False)
    password_hash = db.Column(db.String, nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False)

class Registration(db.Model):
    __tablename__ = "registrations"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=False, unique=True)

    teamname = db.Column(db.String, nullable=False)
    chosen_course = db.Column(db.String, nullable=False)
    diet = db.Column(db.String, nullable=False)

    member1 = db.Column(db.String, nullable=False)
    member2 = db.Column(db.String, nullable=False)

    adress = db.Column(db.String, nullable=False)
    klingelname = db.Column(db.String, nullable=False)

    coordinates = db.Column(db.String, nullable=True)
    special = db.Column(db.String, nullable=True)

    created_at = db.Column(db.DateTime(timezone=True), nullable=False)
    updated_at = db.Column(db.DateTime(timezone=True), nullable=False)

class Assignment(db.Model):
    __tablename__ = "assignments"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=False, unique=True)

    # Storage = db
    route_png_bytes = db.Column(db.LargeBinary, nullable=True)

    # Storage = s3
    route_png_s3_key = db.Column(db.String, nullable=True)

    route_png_filename = db.Column(db.String, nullable=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False)

class Setting(db.Model):
    __tablename__ = "settings"
    key = db.Column(db.String, primary_key=True)
    value = db.Column(db.Text, nullable=False)
    updated_at = db.Column(db.DateTime(timezone=True), nullable=False)

class ChatMessage(db.Model):
    __tablename__ = "chat_messages"
    id = db.Column(db.Integer, primary_key=True)
    nickname = db.Column(db.String, nullable=False)
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False)

with app.app_context():
    db.create_all()

# ---------------- Storage helpers ----------------
def _s3_client():
    import boto3
    kwargs = {}
    if S3_ENDPOINT_URL:
        kwargs["endpoint_url"] = S3_ENDPOINT_URL
    region = (os.environ.get("AWS_REGION") or "").strip()
    if region:
        kwargs["region_name"] = region
    return boto3.client("s3", **kwargs)

def storage_mode() -> str:
    if STORAGE_BACKEND == "s3":
        if not S3_BUCKET:
            raise RuntimeError("STORAGE_BACKEND=s3 but S3_BUCKET is missing.")
        return "s3"
    return "db"

def now_utc():
    return datetime.now(timezone.utc)

def save_png_for_user(user_id: int, filename: str, png_bytes: bytes):
    mode = storage_mode()
    a = db.session.query(Assignment).filter_by(user_id=user_id).first()
    if not a:
        a = Assignment(user_id=user_id, created_at=now_utc())
        db.session.add(a)

    a.route_png_filename = filename

    if mode == "db":
        a.route_png_bytes = png_bytes
        a.route_png_s3_key = None
    else:
        key = f"laufkarten/{user_id}/{secure_filename(filename) or 'laufkarte.png'}"
        s3 = _s3_client()
        s3.put_object(Bucket=S3_BUCKET, Key=key, Body=png_bytes, ContentType="image/png")
        a.route_png_s3_key = key
        a.route_png_bytes = None

    db.session.commit()

def load_png_for_user(user_id: int):
    a = db.session.query(Assignment).filter_by(user_id=user_id).first()
    if not a:
        return None, None

    mode = storage_mode()
    if mode == "db":
        if not a.route_png_bytes:
            return None, a.route_png_filename
        return bytes(a.route_png_bytes), a.route_png_filename
    else:
        if not a.route_png_s3_key:
            return None, a.route_png_filename
        s3 = _s3_client()
        obj = s3.get_object(Bucket=S3_BUCKET, Key=a.route_png_s3_key)
        data = obj["Body"].read()
        return data, a.route_png_filename

# ---------------- General helpers ----------------
@app.context_processor
def inject_globals():
    return {
        "now": lambda: datetime.now(timezone.utc),
        "COURSE_OPTIONS": COURSE_OPTIONS,
        "DIET_OPTIONS": DIET_OPTIONS,
        "APP_NAME": APP_NAME,
        "APP_TAGLINE": APP_TAGLINE,
    }

def _clean(v: str) -> str:
    return (v or "").strip()

def _validate_choice(value: str, allowed, fallback: str) -> str:
    v = (value or "").strip()
    return v if v in allowed else fallback

def _normalize_team_label(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\\s+", " ", s)
    s = re.sub(r"[^\\w\\s\\-äöüß]", "", s, flags=re.UNICODE)
    return s

def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    u = db.session.get(User, int(uid))
    if not u:
        return None
    return {"id": u.id, "email": u.email}

def require_login():
    if not session.get("user_id"):
        flash("Bitte zuerst einloggen.", "warning")
        return redirect(url_for("login"))
    return None

def admin_logged_in():
    return session.get("is_admin") is True

def require_admin():
    if not admin_logged_in():
        flash("Admin-Login erforderlich.", "warning")
        return redirect(url_for("admin_login"))
    return None

def registrations_count() -> int:
    return db.session.query(Registration).count()

def get_setting(key: str, default: str="") -> str:
    row = db.session.get(Setting, key)
    return row.value if row else default

def set_setting(key: str, value: str):
    row = db.session.get(Setting, key)
    if not row:
        row = Setting(key=key, value=value, updated_at=now_utc())
        db.session.add(row)
    else:
        row.value = value
        row.updated_at = now_utc()
    db.session.commit()

def _require_event_pin(pin: str) -> bool:
    if not EVENT_PIN:
        return False
    return _clean(pin) == EVENT_PIN

# ---------------- Routes ----------------
@app.route("/")
def index():
    return render_template(
        "index.html",
        user=current_user(),
        teams_count=registrations_count(),
        event_date=get_setting("event_date", ""),
        info_block=get_setting("info_block", "")
    )

@app.route("/impressum")
def impressum():
    return render_template("impressum.html", user=current_user())

@app.route("/chat", methods=["GET","POST"])
def chat():
    if request.method == "POST":
        nickname = _clean(request.form.get("nickname"))[:40]
        message = _clean(request.form.get("message"))[:500]
        if not nickname or not message:
            flash("Bitte Nickname und Nachricht ausfüllen.", "danger")
            return redirect(url_for("chat"))

        last_ts = session.get("last_chat_ts")
        now_ts = int(datetime.utcnow().timestamp())
        if last_ts and (now_ts - int(last_ts)) < 5:
            flash("Bitte kurz warten (Spam-Schutz).", "warning")
            return redirect(url_for("chat"))
        session["last_chat_ts"] = now_ts

        db.session.add(ChatMessage(nickname=nickname, message=message, created_at=now_utc()))
        db.session.commit()
        flash("Nachricht gesendet.", "success")
        return redirect(url_for("chat"))

    msgs = db.session.query(ChatMessage).order_by(ChatMessage.id.desc()).limit(100).all()
    msgs = list(reversed(msgs))
    msg_rows = [{"nickname": m.nickname, "message": m.message, "created_at": m.created_at.isoformat()} for m in msgs]
    return render_template(
        "chat.html",
        user=current_user(),
        messages=msg_rows,
        teams_count=registrations_count(),
        event_date=get_setting("event_date", ""),
        info_block=get_setting("info_block", "")
    )

@app.route("/register", methods=["GET","POST"])
def register():
    if request.method == "POST":
        pin = request.form.get("event_pin", "")
        if not _require_event_pin(pin):
            flash("Falscher Event-PIN.", "danger")
            return redirect(url_for("register"))

        email = _clean(request.form.get("email", "")).lower()
        password = request.form.get("password", "")

        teamname = _clean(request.form.get("teamname"))
        chosen_course = _validate_choice(request.form.get("chosen_course"), COURSE_OPTIONS, "Egal")
        diet = _validate_choice(request.form.get("diet"), DIET_OPTIONS, "Nichts")

        member1 = _clean(request.form.get("member1"))
        member2 = _clean(request.form.get("member2"))
        adress = _clean(request.form.get("adress"))
        klingelname = _clean(request.form.get("klingelname"))
        coordinates = _clean(request.form.get("coordinates"))
        special = _clean(request.form.get("special"))

        required = [email, password, teamname, chosen_course, diet, member1, member2, adress, klingelname]
        if any(not x for x in required):
            flash("Bitte alle Pflichtfelder ausfüllen.", "danger")
            return redirect(url_for("register"))

        t = now_utc()
        try:
            u = User(email=email, password_hash=generate_password_hash(password), created_at=t)
            db.session.add(u)
            db.session.flush()

            r = Registration(
                user_id=u.id,
                teamname=teamname,
                chosen_course=chosen_course,
                diet=diet,
                member1=member1,
                member2=member2,
                adress=adress,
                klingelname=klingelname,
                coordinates=coordinates or None,
                special=special or None,
                created_at=t,
                updated_at=t,
            )
            db.session.add(r)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Diese E-Mail ist schon registriert. Bitte einloggen.", "warning")
            return redirect(url_for("login"))

        session["user_id"] = u.id
        flash("Registrierung erfolgreich.", "success")
        return redirect(url_for("dashboard"))

    return render_template(
        "register.html",
        user=current_user(),
        teams_count=registrations_count(),
        event_date=get_setting("event_date", ""),
        info_block=get_setting("info_block", "")
    )

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        email = _clean(request.form.get("email", "")).lower()
        password = request.form.get("password", "")

        u = db.session.query(User).filter_by(email=email).first()
        if not u or not check_password_hash(u.password_hash, password):
            flash("Login fehlgeschlagen.", "danger")
            return redirect(url_for("login"))

        session["user_id"] = u.id
        flash("Eingeloggt.", "success")
        return redirect(url_for("dashboard"))

    return render_template(
        "login.html",
        user=current_user(),
        teams_count=registrations_count(),
        event_date=get_setting("event_date", ""),
        info_block=get_setting("info_block", "")
    )

@app.route("/logout")
def logout():
    session.pop("user_id", None)
    flash("Ausgeloggt.", "info")
    return redirect(url_for("index"))

@app.route("/dashboard", methods=["GET","POST"])
def dashboard():
    need = require_login()
    if need:
        return need

    u = db.session.get(User, int(session["user_id"]))
    reg = db.session.query(Registration).filter_by(user_id=u.id).first()
    assignment = db.session.query(Assignment).filter_by(user_id=u.id).first()

    if request.method == "POST":
        teamname = _clean(request.form.get("teamname"))
        chosen_course = _validate_choice(request.form.get("chosen_course"), COURSE_OPTIONS, "Egal")
        diet = _validate_choice(request.form.get("diet"), DIET_OPTIONS, "Nichts")
        member1 = _clean(request.form.get("member1"))
        member2 = _clean(request.form.get("member2"))
        adress = _clean(request.form.get("adress"))
        klingelname = _clean(request.form.get("klingelname"))
        coordinates = _clean(request.form.get("coordinates"))
        special = _clean(request.form.get("special"))

        required = [teamname, chosen_course, diet, member1, member2, adress, klingelname]
        if any(not x for x in required):
            flash("Bitte alle Pflichtfelder ausfüllen.", "danger")
            return redirect(url_for("dashboard"))

        t = now_utc()
        if not reg:
            reg = Registration(
                user_id=u.id,
                teamname=teamname, chosen_course=chosen_course, diet=diet,
                member1=member1, member2=member2,
                adress=adress, klingelname=klingelname,
                coordinates=coordinates or None, special=special or None,
                created_at=t, updated_at=t
            )
            db.session.add(reg)
        else:
            reg.teamname = teamname
            reg.chosen_course = chosen_course
            reg.diet = diet
            reg.member1 = member1
            reg.member2 = member2
            reg.adress = adress
            reg.klingelname = klingelname
            reg.coordinates = coordinates or None
            reg.special = special or None
            reg.updated_at = t

        db.session.commit()
        flash("Daten gespeichert.", "success")
        return redirect(url_for("dashboard"))

    reg_dict = None
    if reg:
        reg_dict = {
            "teamname": reg.teamname,
            "chosen_course": reg.chosen_course,
            "diet": reg.diet,
            "Member1": reg.member1,
            "Member2": reg.member2,
            "Adress": reg.adress,
            "Klingelname": reg.klingelname,
            "Coordinates": reg.coordinates,
            "Special": reg.special,
        }
    asg_dict = None
    if assignment and (assignment.route_png_bytes or assignment.route_png_s3_key):
        asg_dict = {"route_png_path": "stored", "route_png_filename": assignment.route_png_filename}

    return render_template(
        "dashboard.html",
        user={"id": u.id, "email": u.email},
        reg=reg_dict,
        assignment=asg_dict,
        teams_count=registrations_count(),
        event_date=get_setting("event_date", ""),
        info_block=get_setting("info_block", "")
    )

@app.route("/dashboard/laufkarte.png")
def dashboard_laufkarte_png():
    need = require_login()
    if need:
        return need
    user_id = int(session["user_id"])
    data, _fn = load_png_for_user(user_id)
    if not data:
        abort(404)
    return send_file(io.BytesIO(data), mimetype="image/png", as_attachment=False)

@app.route("/dashboard/laufkarte.png/download")
def dashboard_laufkarte_png_download():
    need = require_login()
    if need:
        return need
    user_id = int(session["user_id"])
    data, fn = load_png_for_user(user_id)
    if not data:
        abort(404)
    return send_file(io.BytesIO(data), mimetype="image/png", as_attachment=True, download_name=(fn or "laufkarte.png"))

@app.route("/admin")
def admin_root():
    if not admin_logged_in():
        return redirect(url_for("admin_login"))
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    if request.method == "POST":
        pw = request.form.get("password", "")
        if pw == ADMIN_PASSWORD:
            session["is_admin"] = True
            flash("Admin eingeloggt.", "success")
            return redirect(url_for("admin_dashboard"))
        flash("Falsches Admin-Passwort.", "danger")
        return redirect(url_for("admin_login"))

    return render_template(
        "admin_login.html",
        user=current_user(),
        teams_count=registrations_count(),
        event_date=get_setting("event_date", ""),
        info_block=get_setting("info_block", "")
    )

@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    flash("Admin ausgeloggt.", "info")
    return redirect(url_for("index"))

@app.route("/admin/dashboard", methods=["GET","POST"])
def admin_dashboard():
    need = require_admin()
    if need:
        return need

    if request.method == "POST":
        set_setting("event_date", _clean(request.form.get("event_date")))
        set_setting("info_block", _clean(request.form.get("info_block")))
        flash("Event-Daten gespeichert.", "success")
        return redirect(url_for("admin_dashboard"))

    rows = (db.session.query(User.email,
                             Registration.teamname,
                             Registration.chosen_course,
                             Registration.diet,
                             Assignment.route_png_filename)
            .outerjoin(Registration, Registration.user_id == User.id)
            .outerjoin(Assignment, Assignment.user_id == User.id)
            .order_by(Registration.created_at.desc().nullslast())
            .all())
    out_rows = [{"email": r[0], "teamname": r[1], "chosen_course": r[2], "diet": r[3], "route_png_filename": r[4]} for r in rows]

    return render_template(
        "admin_dashboard.html",
        rows=out_rows,
        user=current_user(),
        teams_count=registrations_count(),
        event_date=get_setting("event_date", ""),
        info_block=get_setting("info_block", "")
    )

@app.route("/admin/export.xlsx")
def admin_export_xlsx():
    need = require_admin()
    if need:
        return need

    regs = db.session.query(Registration).order_by(Registration.created_at.desc()).all()

    if os.path.exists(XLSX_TEMPLATE_PATH):
        wb = openpyxl.load_workbook(XLSX_TEMPLATE_PATH)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Number", "teamname", "chosen course", "diet", "Member1", "Member2", "Adress", "Klingelname", "Coordinates", "Special"])

    for i, r in enumerate(regs, start=1):
        ws.append([i, r.teamname, r.chosen_course, r.diet, r.member1, r.member2, r.adress, r.klingelname, r.coordinates or "", r.special or ""])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="registrations_for_matching.xlsx")

@app.route("/admin/upload-laufkarten", methods=["GET","POST"])
def admin_upload_laufkarten():
    need = require_admin()
    if need:
        return need

    if request.method == "POST":
        zfile = request.files.get("zipfile")
        if not zfile or zfile.filename == "":
            flash("Bitte ZIP-Datei auswählen.", "danger")
            return redirect(url_for("admin_upload_laufkarten"))

        raw = zfile.read()
        try:
            zf = zipfile.ZipFile(io.BytesIO(raw))
        except Exception:
            flash("ZIP-Datei konnte nicht gelesen werden.", "danger")
            return redirect(url_for("admin_upload_laufkarten"))

        regs = db.session.query(Registration.user_id, Registration.teamname).all()
        team_index = {_normalize_team_label(t): (uid, t) for (uid, t) in regs}

        imported = 0
        skipped = []

        for info in zf.infolist():
            if info.is_dir():
                continue
            name = info.filename
            if "__MACOSX" in name or name.endswith("/"):
                continue
            if not name.lower().endswith(".png"):
                continue

            stem = os.path.basename(name)[:-4]
            key = _normalize_team_label(stem)
            if key not in team_index:
                skipped.append(stem)
                continue

            user_id, teamname = team_index[key]
            png_bytes = zf.read(info)
            safe_name = secure_filename(f"{teamname}.png") or f"user_{user_id}.png"
            save_png_for_user(int(user_id), safe_name, png_bytes)
            imported += 1

        msg = f"PNGs importiert: {imported} (Storage: {storage_mode()})."
        if skipped:
            msg += " Nicht gematchte Dateien: " + ", ".join(skipped[:10])
            if len(skipped) > 10:
                msg += f" (+{len(skipped)-10} weitere)"
        flash(msg, "success")
        return redirect(url_for("admin_dashboard"))

    return render_template(
        "admin_upload_laufkarten.html",
        user=current_user(),
        teams_count=registrations_count(),
        event_date=get_setting("event_date", ""),
        info_block=get_setting("info_block", "")
    )

if __name__ == "__main__":
    app.run(debug=True)
